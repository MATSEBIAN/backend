"""
Matsebian ERP v2.0 - Backend API
Flask + SQLite (puro, sin SQLAlchemy)
Solo requiere: flask, openpyxl, werkzeug (ya instalados)
"""
import os, re, json, base64, sqlite3, traceback, urllib.request, urllib.error
from datetime import datetime
import jwt as pyjwt
import datetime
from functools import wraps
from flask import Flask, request, jsonify, session, send_from_directory, g
from werkzeug.security import generate_password_hash, check_password_hash

ANTHROPIC_API_KEY = os.environ.get('ANTHROPIC_API_KEY', '')

app = Flask(__name__, static_folder='frontend', static_url_path='')
app.config['SESSION_COOKIE_SAMESITE'] = 'None'
app.config['SESSION_COOKIE_SECURE'] = True
JWT_SECRET = os.environ.get('JWT_SECRET', 'matsebian-jwt-secret-2026')
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'matsebian-erp-dev-2026')
app.config['DATABASE'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'matsebian_erp.db')
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# === DB Helpers ===
def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(app.config['DATABASE'])
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA foreign_keys=ON")
    return g.db

@app.teardown_appcontext
def close_db(exc):
    db = g.pop('db', None)
    if db: db.close()

def qry(sql, args=(), one=False):
    cur = get_db().execute(sql, args)
    rv = [dict(row) for row in cur.fetchall()]
    return rv[0] if one and rv else (None if one else rv)

def exe(sql, args=()):
    db = get_db(); cur = db.execute(sql, args); db.commit(); return cur.lastrowid

# === CORS manual ===
@app.after_request
def cors(r):
    r.headers['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
    r.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    r.headers['Access-Control-Allow-Methods'] = 'GET,POST,PUT,DELETE,OPTIONS'
    r.headers['Access-Control-Allow-Credentials'] = 'true'
    return r

@app.route('/api/<path:p>', methods=['OPTIONS'])
def opts(p): return '', 204

# === Auth ===
def login_required(f):
    @wraps(f)
    def d(*a, **k):
        auth = request.headers.get('Authorization', '')
        if auth.startswith('Bearer '):
            try:
                payload = pyjwt.decode(auth[7:], JWT_SECRET, algorithms=['HS256'])
                session['user_id'] = payload['user_id']
            except Exception:
                return jsonify({'error': 'Token invalido'}), 401
        elif 'user_id' not in session:
            return jsonify({'msg': 'Missing Authorization Header'}), 401
        return f(*a, **k)
    return d

def get_user_empresas(uid):
    empresas = qry('SELECT e.*, ue.permisos FROM empresas e JOIN usuarios_empresas ue ON e.id=ue.empresa_id WHERE ue.usuario_id=? AND e.activo=1', [uid])
    for e in empresas:
        e['locales'] = qry('SELECT * FROM locales WHERE empresa_id=? AND activo=1', [e['id']])
    return empresas

@app.route('/api/auth/login', methods=['POST'])
def login():
    d = request.get_json()
    u = qry('SELECT * FROM usuarios WHERE email=? AND activo=1', [d.get('email','').strip().lower()], one=True)
    if not u or not check_password_hash(u['password_hash'], d.get('password','')):
        return jsonify({'error': 'Credenciales incorrectas'}), 401
    session['user_id'] = u['id']
    token = pyjwt.encode({'user_id': u['id'], 'exp': datetime.datetime.utcnow() + datetime.timedelta(days=30)}, JWT_SECRET, algorithm='HS256')
    return jsonify({'user': {k:u[k] for k in ['id','email','nombre','rol']}, 'empresas': get_user_empresas(u['id']), 'token': token, 'access_token': token, 'company': get_user_empresas(u['id'])[0] if get_user_empresas(u['id']) else None})

@app.route('/api/auth/logout', methods=['POST'])
def logout(): session.clear(); return jsonify({'ok': True})

@app.route('/api/auth/me')
@login_required
def me():
    u = qry('SELECT * FROM usuarios WHERE id=?', [session['user_id']], one=True)
    if not u: return jsonify({'error': 'No encontrado'}), 401
    return jsonify({'user': {k:u[k] for k in ['id','email','nombre','rol']}, 'empresas': get_user_empresas(u['id'])})

# === Empresas ===
@app.route('/api/empresas')
@login_required
def list_empresas(): return jsonify(get_user_empresas(session['user_id']))

@app.route('/api/empresas', methods=['POST'])
@login_required
def create_empresa():
    d = request.get_json()
    eid = exe('INSERT INTO empresas(nombre,nombre_corto,cif,tipo,color) VALUES(?,?,?,?,?)',
              [d['nombre'], d.get('nombre_corto',''), d.get('cif'), d.get('tipo','restaurante'), d.get('color','#00ff88')])
    exe('INSERT INTO usuarios_empresas(usuario_id,empresa_id,permisos) VALUES(?,?,?)', [session['user_id'], eid, 'admin'])
    for loc in d.get('locales', []):
        exe('INSERT INTO locales(empresa_id,nombre,nombre_corto,ciudad) VALUES(?,?,?,?)',
            [eid, loc['nombre'], loc.get('nombre_corto'), loc.get('ciudad')])
    return jsonify(qry('SELECT * FROM empresas WHERE id=?', [eid], one=True)), 201

# === Locales ===
@app.route('/api/empresas/<int:eid>/locales')
@login_required
def list_locales(eid): return jsonify(qry('SELECT * FROM locales WHERE empresa_id=? AND activo=1', [eid]))

@app.route('/api/empresas/<int:eid>/locales', methods=['POST'])
@login_required
def create_local(eid):
    d = request.get_json()
    lid = exe('INSERT INTO locales(empresa_id,nombre,nombre_corto,ciudad) VALUES(?,?,?,?)',
              [eid, d['nombre'], d.get('nombre_corto'), d.get('ciudad')])
    return jsonify(qry('SELECT * FROM locales WHERE id=?', [lid], one=True)), 201

# === Facturas ===
@app.route('/api/empresas/<int:eid>/facturas')
@login_required
def list_facturas(eid):
    sql = 'SELECT f.*, l.nombre as local_nombre FROM facturas f LEFT JOIN locales l ON f.local_id=l.id WHERE f.empresa_id=?'
    p = [eid]
    if request.args.get('year'): sql += ' AND strftime("%%Y",f.fecha)=?'; p.append(request.args['year'])
    if request.args.get('month'): sql += ' AND strftime("%%Y-%%m",f.fecha)=?'; p.append(request.args['month'])
    if request.args.get('local_id'): sql += ' AND f.local_id=?'; p.append(int(request.args['local_id']))
    return jsonify(qry(sql + ' ORDER BY f.fecha DESC', p))

@app.route('/api/empresas/<int:eid>/facturas', methods=['POST'])
@login_required
def create_factura(eid):
    d = request.get_json()
    local_id = d.get('local_id')
    if not local_id and d.get('local_nombre'):
        n = d['local_nombre'].upper()
        for l in qry('SELECT * FROM locales WHERE empresa_id=? AND activo=1', [eid]):
            if l['nombre'].upper() in n or n in l['nombre'].upper() or (l['nombre_corto'] and l['nombre_corto'].upper() in n):
                local_id = l['id']; break
    fid = exe('INSERT INTO facturas(empresa_id,local_id,fecha,num_factura,proveedor,cif_proveedor,concepto,base,iva,irpf,total,origen) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)',
              [eid, local_id, d.get('fecha',datetime.now().strftime('%Y-%m-%d')), d.get('num_factura'), d.get('proveedor','?'),
               d.get('cif_proveedor'), d.get('concepto'), float(d.get('base',0)), float(d.get('iva',0)),
               float(d.get('irpf',0)), float(d.get('total',0)), d.get('origen','manual')])
    return jsonify(qry('SELECT * FROM facturas WHERE id=?', [fid], one=True)), 201

@app.route('/api/empresas/<int:eid>/facturas/<int:fid>', methods=['DELETE'])
@login_required
def delete_factura(eid, fid):
    exe('DELETE FROM facturas WHERE id=? AND empresa_id=?', [fid, eid]); return jsonify({'ok': True})

# === Import Excel ===
@app.route('/api/empresas/<int:eid>/importar-excel', methods=['POST'])
@login_required
def importar_excel(eid):
    if 'file' not in request.files: return jsonify({'error': 'No file'}), 400
    f = request.files['file']
    try:
        import openpyxl; from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(f.read()), data_only=True); ws = wb.active
        hdrs = [str(c.value or '').strip() for c in ws[1]]
        aliases = {'FECHA':['FECHA','Fecha'],'ACREEDOR':['ACREEDOR','Acreedor','PROVEEDOR'],'P&L':['P&L','P and L','CONCEPTO'],
                   'BASE':['BASE','Base'],'IVA':['IVA'],'TOTAL':['TOTAL FACT','TOTAL'],'IRPF':['ALBARAN/IRPF','IRPF','ALBARAN'],
                   'CIF':['CIF','NIF'],'NFACT':['Nº FACTURA','NUM FACTURA'],'LOCAL':['LOCAL','Local']}
        ci = {}
        for k, ns in aliases.items():
            for i, h in enumerate(hdrs):
                if h in ns: ci[k]=i; break
        locales = qry('SELECT * FROM locales WHERE empresa_id=? AND activo=1', [eid])
        db = get_db(); imp = 0; errs = []
        for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                if not row or all(v is None for v in row): continue
                def v(k, d=None):
                    i=ci.get(k); return row[i] if i is not None and i<len(row) and row[i] is not None else d
                prov = str(v('ACREEDOR') or '').strip()
                if not prov: continue
                fr = v('FECHA')
                if isinstance(fr,(datetime,date)): fecha=fr.strftime('%Y-%m-%d') if isinstance(fr,datetime) else fr.isoformat()
                elif isinstance(fr,str):
                    fecha=None
                    for fmt in ['%Y-%m-%d','%d/%m/%Y','%d-%m-%Y']:
                        try: fecha=datetime.strptime(fr.strip(),fmt).strftime('%Y-%m-%d'); break
                        except: pass
                    if not fecha: fecha=datetime.now().strftime('%Y-%m-%d')
                else: fecha=datetime.now().strftime('%Y-%m-%d')
                loc_raw = str(v('LOCAL') or '').strip().upper(); lid=None
                for l in locales:
                    if l['nombre'].upper() in loc_raw or loc_raw in l['nombre'].upper() or (l['nombre_corto'] and l['nombre_corto'].upper() in loc_raw):
                        lid=l['id']; break
                db.execute('INSERT INTO facturas(empresa_id,local_id,fecha,num_factura,proveedor,cif_proveedor,concepto,base,iva,irpf,total,origen) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)',
                    [eid, lid, fecha, str(v('NFACT') or '') or None, prov, str(v('CIF') or '') or None,
                     str(v('P&L') or ''), float(v('BASE') or 0), float(v('IVA') or 0), float(v('IRPF') or 0), float(v('TOTAL') or 0), 'excel'])
                imp += 1
            except Exception as e: errs.append(f'Fila {ri}: {e}')
        db.commit()
        return jsonify({'ok':True, 'imported':imp, 'errors':errs[:10]})
    except Exception as e: traceback.print_exc(); return jsonify({'error':str(e)}), 500

# === Ventas ===
@app.route('/api/empresas/<int:eid>/ventas')
@login_required
def list_ventas(eid):
    sql = 'SELECT v.*, l.nombre as local_nombre FROM ventas_periodo v LEFT JOIN locales l ON v.local_id=l.id WHERE v.empresa_id=?'
    p = [eid]
    if request.args.get('periodo'): sql += ' AND v.periodo=?'; p.append(request.args['periodo'])
    return jsonify(qry(sql + ' ORDER BY v.periodo DESC', p))

@app.route('/api/empresas/<int:eid>/ventas', methods=['POST'])
@login_required
def upsert_ventas(eid):
    d = request.get_json(); per=d['periodo']; lid=d.get('local_id')
    ex = qry('SELECT id FROM ventas_periodo WHERE empresa_id=? AND periodo=? AND (local_id=? OR (local_id IS NULL AND ? IS NULL))', [eid,per,lid,lid], one=True)
    if ex: exe('UPDATE ventas_periodo SET ventas_total=?, coste_laboral=? WHERE id=?', [float(d.get('ventas_total',0)), float(d.get('coste_laboral',0)), ex['id']]); vid=ex['id']
    else: vid = exe('INSERT INTO ventas_periodo(empresa_id,local_id,periodo,ventas_total,coste_laboral) VALUES(?,?,?,?,?)', [eid,lid,per,float(d.get('ventas_total',0)),float(d.get('coste_laboral',0))])
    return jsonify(qry('SELECT * FROM ventas_periodo WHERE id=?', [vid], one=True))

# === OCR ===
@app.route('/api/empresas/<int:eid>/ocr', methods=['POST'])
@login_required
def ocr_factura(eid):
    if not ANTHROPIC_API_KEY:
        return jsonify({'error': 'Configura ANTHROPIC_API_KEY. Ejecuta: export ANTHROPIC_API_KEY=tu-clave-aqui'}), 400
    d = request.get_json()
    img = d.get('image','')
    if not img: return jsonify({'error': 'No image'}), 400
    if ',' in img: img = img.split(',')[1]
    locales = qry('SELECT * FROM locales WHERE empresa_id=? AND activo=1', [eid])
    loc_text = ', '.join([f"{l['nombre']} ({l['nombre_corto']})" if l['nombre_corto'] else l['nombre'] for l in locales])
    body = json.dumps({
        "model": "claude-sonnet-4-20250514", "max_tokens": 2000,
        "messages": [{"role":"user","content":[
            {"type":"image","source":{"type":"base64","media_type":"image/jpeg","data":img}},
            {"type":"text","text":f"""Analiza esta factura y extrae los datos en JSON.
Locales disponibles: {loc_text}
IMPORTANTE: Prioriza "ENVIADO A"/"SHIP TO" sobre "FACTURADO A" para el local.
Responde SOLO JSON valido:
{{"fecha":"YYYY-MM-DD","proveedor":"nombre","cif_proveedor":"CIF","num_factura":"numero",
"concepto":"categoria P&L","base":0.00,"iva":0.00,"irpf":0.00,"total":0.00,
"local_nombre":"local detectado o null","notas":"observaciones"}}"""}
        ]}]
    }).encode('utf-8')
    try:
        req = urllib.request.Request('https://api.anthropic.com/v1/messages', data=body,
            headers={'Content-Type':'application/json','x-api-key':ANTHROPIC_API_KEY,'anthropic-version':'2023-06-01'})
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json.loads(resp.read().decode('utf-8'))
        txt = result['content'][0]['text']
        if '```json' in txt: txt = txt.split('```json')[1].split('```')[0]
        elif '```' in txt: txt = txt.split('```')[1].split('```')[0]
        ocr = json.loads(txt.strip())
        # Save image
        fn = f"ocr_{eid}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
        with open(os.path.join(app.config['UPLOAD_FOLDER'], fn), 'wb') as f:
            f.write(base64.b64decode(img))
        ocr['documento_path'] = fn; ocr['origen'] = 'ocr'
        return jsonify(ocr)
    except urllib.error.HTTPError as e:
        err = e.read().decode('utf-8') if e.readable() else str(e)
        return jsonify({'error': f'API Anthropic: {err}'}), 500
    except Exception as e:
        traceback.print_exc(); return jsonify({'error': str(e)}), 500

# === API Key config ===
@app.route('/api/config/apikey', methods=['POST'])
@login_required
def set_apikey():
    global ANTHROPIC_API_KEY
    d = request.get_json()
    key = d.get('key','').strip()
    if not key.startswith('sk-'): return jsonify({'error': 'La clave debe empezar con sk-'}), 400
    ANTHROPIC_API_KEY = key
    return jsonify({'ok': True})

# === Dashboard ===
@app.route('/api/empresas/<int:eid>/dashboard')
@login_required
def dashboard(eid):
    w=['f.empresa_id=?']; p=[eid]
    if request.args.get('year'): w.append('strftime("%%Y",f.fecha)=?'); p.append(request.args['year'])
    if request.args.get('month'): w.append('strftime("%%Y-%%m",f.fecha)=?'); p.append(request.args['month'])
    if request.args.get('local_id'): w.append('f.local_id=?'); p.append(int(request.args['local_id']))
    ws = ' AND '.join(w)
    s = qry(f'SELECT COUNT(*) as tf, COALESCE(SUM(base),0) as tb, COALESCE(SUM(iva),0) as ti, COALESCE(SUM(total),0) as tg, COALESCE(SUM(irpf),0) as tr FROM facturas f WHERE {ws}', p, one=True)
    # Ventas
    vw=['v.empresa_id=?']; vp=[eid]
    if request.args.get('month'): vw.append('v.periodo=?'); vp.append(request.args['month'])
    elif request.args.get('year'): vw.append('v.periodo LIKE ?'); vp.append(f"{request.args['year']}-%")
    if request.args.get('local_id'): vw.append('v.local_id=?'); vp.append(int(request.args['local_id']))
    vd = qry(f"SELECT v.*, l.nombre as ln FROM ventas_periodo v LEFT JOIN locales l ON v.local_id=l.id WHERE {' AND '.join(vw)}", vp)
    tv=sum(v['ventas_total'] for v in vd); tc=sum(v['coste_laboral'] for v in vd)
    ga=abs(s['tg']); rc=tv-tc-ga
    iv=tv-(tv/1.1) if tv>0 else 0; ig=abs(s['ti']); ip=iv-ig; cf=rc-s['tr']-(ip if ip>0 else 0)
    vs=tv/1.1 if tv>0 else 0; ba=abs(s['tb']); rf=vs-tc-ba
    return jsonify({
        'resumen': {'total_facturas':s['tf'],'total_base':s['tb'],'total_iva':s['ti'],'total_gastado':s['tg'],'total_irpf':s['tr'],'promedio_factura':s['tg']/s['tf'] if s['tf']>0 else 0},
        'ventas': {'total_ventas':tv,'total_coste_laboral':tc,'por_local':{v['ln'] or 'General':v['ventas_total'] for v in vd},'coste_por_local':{v['ln'] or 'General':v['coste_laboral'] for v in vd}},
        'cashflow': {'ventas':tv,'coste_laboral':tc,'total_gastado':ga,'resultado':rc,'irpf':s['tr'],'iva_pagar':ip,'cash_flow':cf},
        'fiscal': {'ventas_sin_iva':vs,'coste_laboral':tc,'total_gastado_base':ba,'resultado':rf},
        'graficos': {
            'por_proveedor': qry(f'SELECT proveedor as nombre, SUM(ABS(total)) as total FROM facturas f WHERE {ws} GROUP BY proveedor ORDER BY total DESC LIMIT 10', p),
            'por_categoria': qry(f'SELECT COALESCE(concepto,"Sin cat") as nombre, SUM(ABS(total)) as total FROM facturas f WHERE {ws} GROUP BY concepto ORDER BY total DESC LIMIT 10', p),
            'por_mes': qry(f'SELECT strftime("%%Y-%%m",fecha) as mes, SUM(ABS(total)) as total FROM facturas f WHERE {ws} GROUP BY mes ORDER BY mes', p),
            'por_local': qry(f'SELECT COALESCE(l.nombre,"Sin local") as nombre, SUM(ABS(f.total)) as total FROM facturas f LEFT JOIN locales l ON f.local_id=l.id WHERE {ws} GROUP BY l.nombre ORDER BY total DESC', p),
        },
        'filtros_disponibles': {
            'years': [r['y'] for r in qry('SELECT DISTINCT strftime("%%Y",fecha) as y FROM facturas WHERE empresa_id=? ORDER BY y', [eid])],
            'months': [r['m'] for r in qry('SELECT DISTINCT strftime("%%Y-%%m",fecha) as m FROM facturas WHERE empresa_id=? ORDER BY m', [eid])],
            'locales': qry('SELECT id, nombre FROM locales WHERE empresa_id=? AND activo=1', [eid])
        }
    })

# === Frontend ===
@app.route('/')
def index(): return send_from_directory(app.static_folder, 'index.html')

# === Init DB ===
def init_db():
    db = sqlite3.connect(app.config['DATABASE'])
    db.executescript('''
        CREATE TABLE IF NOT EXISTS usuarios(id INTEGER PRIMARY KEY AUTOINCREMENT, email TEXT UNIQUE NOT NULL, password_hash TEXT NOT NULL, nombre TEXT NOT NULL, rol TEXT DEFAULT 'admin', activo INTEGER DEFAULT 1, created_at TEXT DEFAULT(datetime('now')));
        CREATE TABLE IF NOT EXISTS empresas(id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL, nombre_corto TEXT NOT NULL, cif TEXT, tipo TEXT DEFAULT 'restaurante', moneda TEXT DEFAULT 'EUR', iva_por_defecto REAL DEFAULT 10.0, color TEXT DEFAULT '#00ff88', activo INTEGER DEFAULT 1, created_at TEXT DEFAULT(datetime('now')));
        CREATE TABLE IF NOT EXISTS locales(id INTEGER PRIMARY KEY AUTOINCREMENT, empresa_id INTEGER NOT NULL REFERENCES empresas(id), nombre TEXT NOT NULL, nombre_corto TEXT, direccion TEXT, ciudad TEXT, activo INTEGER DEFAULT 1, created_at TEXT DEFAULT(datetime('now')));
        CREATE TABLE IF NOT EXISTS categorias_gasto(id INTEGER PRIMARY KEY AUTOINCREMENT, empresa_id INTEGER NOT NULL REFERENCES empresas(id), codigo TEXT NOT NULL, nombre TEXT NOT NULL, grupo TEXT, activo INTEGER DEFAULT 1);
        CREATE TABLE IF NOT EXISTS facturas(id INTEGER PRIMARY KEY AUTOINCREMENT, empresa_id INTEGER NOT NULL REFERENCES empresas(id), local_id INTEGER REFERENCES locales(id), categoria_id INTEGER REFERENCES categorias_gasto(id), fecha TEXT NOT NULL, num_factura TEXT, proveedor TEXT NOT NULL, cif_proveedor TEXT, concepto TEXT, base REAL DEFAULT 0, iva REAL DEFAULT 0, irpf REAL DEFAULT 0, total REAL DEFAULT 0, notas TEXT, documento_path TEXT, origen TEXT DEFAULT 'manual', created_at TEXT DEFAULT(datetime('now')), updated_at TEXT DEFAULT(datetime('now')));
        CREATE TABLE IF NOT EXISTS ventas_periodo(id INTEGER PRIMARY KEY AUTOINCREMENT, empresa_id INTEGER NOT NULL REFERENCES empresas(id), local_id INTEGER REFERENCES locales(id), periodo TEXT NOT NULL, ventas_total REAL DEFAULT 0, coste_laboral REAL DEFAULT 0, notas TEXT, created_at TEXT DEFAULT(datetime('now')), UNIQUE(empresa_id,local_id,periodo));
        CREATE TABLE IF NOT EXISTS usuarios_empresas(id INTEGER PRIMARY KEY AUTOINCREMENT, usuario_id INTEGER NOT NULL REFERENCES usuarios(id), empresa_id INTEGER NOT NULL REFERENCES empresas(id), permisos TEXT DEFAULT 'admin', UNIQUE(usuario_id,empresa_id));
        CREATE INDEX IF NOT EXISTS idx_f_emp ON facturas(empresa_id);
        CREATE INDEX IF NOT EXISTS idx_f_fecha ON facturas(fecha);
        CREATE INDEX IF NOT EXISTS idx_f_local ON facturas(local_id);
        CREATE TABLE IF NOT EXISTS transactions(id INTEGER PRIMARY KEY AUTOINCREMENT, empresa_id INTEGER NOT NULL REFERENCES empresas(id), type TEXT NOT NULL DEFAULT 'expense', amount REAL NOT NULL DEFAULT 0, description TEXT, payment_method TEXT DEFAULT 'cash', transaction_date TEXT NOT NULL, category_id INTEGER, vendor_client TEXT, tax_amount REAL DEFAULT 0, notes TEXT, source TEXT DEFAULT 'manual', created_at TEXT DEFAULT(datetime('now')));
        CREATE TABLE IF NOT EXISTS reports (id INTEGER PRIMARY KEY AUTOINCREMENT, empresa_id INTEGER, year INTEGER, month INTEGER, content TEXT, created_at TEXT);
        CREATE TABLE IF NOT EXISTS transaction_categories(id INTEGER PRIMARY KEY AUTOINCREMENT, empresa_id INTEGER NOT NULL REFERENCES empresas(id), name TEXT NOT NULL, type TEXT DEFAULT 'both', activo INTEGER DEFAULT 1);
    ''')
    if db.execute('SELECT COUNT(*) FROM usuarios').fetchone()[0] == 0:
        pw = generate_password_hash('admin123')
        db.execute('INSERT INTO usuarios(email,password_hash,nombre,rol) VALUES(?,?,?,?)', ('daniel@matsebian.com',pw,'Daniel','admin'))
        db.execute("INSERT INTO empresas(nombre,nombre_corto,cif,tipo,color) VALUES(?,?,?,?,?)", ('Las Adelitas - Sabores Adelita S.L.','adelitas','B12345678','restaurante','#00ff88'))
        db.execute("INSERT INTO empresas(nombre,nombre_corto,tipo,color) VALUES(?,?,?,?)", ("Carl's Jr - Morenlonia S.L.",'carlsjr','franquicia','#ff6600'))
        db.execute("INSERT INTO locales(empresa_id,nombre,nombre_corto,ciudad) VALUES(1,'Las Adelitas Madrid','LAD','Madrid')")
        db.execute("INSERT INTO locales(empresa_id,nombre,nombre_corto,ciudad) VALUES(2,'Dos Hermanas','WAY','Dos Hermanas')")
        db.execute("INSERT INTO locales(empresa_id,nombre,nombre_corto,ciudad) VALUES(2,'Jerez','LUZ','Jerez de la Frontera')")
        db.execute("INSERT INTO usuarios_empresas(usuario_id,empresa_id,permisos) VALUES(1,1,'admin')")
        db.execute("INSERT INTO usuarios_empresas(usuario_id,empresa_id,permisos) VALUES(1,2,'admin')")
        db.commit(); print('✅ DB initialized with seed data')
    db.close()


# TRANSACTIONS API
def get_first_empresa(uid):
    e = qry('SELECT empresa_id FROM usuarios_empresas WHERE usuario_id=? LIMIT 1', [uid], one=True)
    return e['empresa_id'] if e else None

@app.route('/api/transactions/')
@login_required
def list_transactions():
    eid = get_first_empresa(session['user_id'])
    if not eid: return jsonify([])
    d = request.args
    sql = "SELECT t.*, tc.name as category FROM transactions t LEFT JOIN transaction_categories tc ON t.category_id=tc.id WHERE t.empresa_id=?"
    args = [eid]
    if d.get('type'):  sql += ' AND t.type=?';  args.append(d['type'])
    if d.get('month'): sql += " AND strftime('%m', t.transaction_date)=?"; args.append(str(d['month']).zfill(2))
    if d.get('year'):  sql += " AND strftime('%Y', t.transaction_date)=?"; args.append(str(d['year']))
    sql += ' ORDER BY t.transaction_date DESC, t.id DESC'
    return jsonify(qry(sql, args))

@app.route('/api/transactions/manual', methods=['POST'])
@login_required
def create_transaction():
    eid = get_first_empresa(session['user_id'])
    if not eid: return jsonify({'error': 'Sin empresa'}), 400
    d = request.get_json()
    if not d or not d.get('amount') or not d.get('description'):
        return jsonify({'error': 'Faltan campos obligatorios'}), 400
    exe("""INSERT INTO transactions(empresa_id,type,amount,description,payment_method,transaction_date,category_id,vendor_client,tax_amount,notes,source) VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
        [eid, d.get('type','expense'), float(d['amount']), d['description'],
         d.get('payment_method','cash'), d.get('transaction_date', str(datetime.date.today())),
         d.get('category_id'), d.get('vendor_client',''), float(d.get('tax_amount') or 0),
         d.get('notes',''), 'manual'])
    return jsonify({'ok': True}), 201

@app.route('/api/transactions/<int:tid>', methods=['DELETE'])
@login_required
def delete_transaction(tid):
    eid = get_first_empresa(session['user_id'])
    exe('DELETE FROM transactions WHERE id=? AND empresa_id=?', [tid, eid])
    return jsonify({'ok': True})

@app.route('/api/transactions/categories')
@login_required
def list_tx_categories():
    eid = get_first_empresa(session['user_id'])
    return jsonify(qry('SELECT * FROM transaction_categories WHERE empresa_id=? AND activo=1', [eid]))

@app.route('/api/transactions/categories', methods=['POST'])
@login_required
def create_tx_category():
    eid = get_first_empresa(session['user_id'])
    d = request.get_json()
    exe('INSERT INTO transaction_categories(empresa_id,name,type) VALUES(?,?,?)',
        [eid, d['name'], d.get('type','both')])
    return jsonify({'ok': True}), 201

@app.route('/api/dashboard/')
@login_required
def dashboard_frontend():
    eid = get_first_empresa(session['user_id'])
    if not eid: return jsonify({'ingresos':0,'gastos':0,'neto':0,'margen':0,'por_categoria':[],'ultimos':[]})
    month = request.args.get('month')
    year  = request.args.get('year')
    w = ['empresa_id=?']; p = [eid]
    if year:  w.append('strftime("%Y",transaction_date)=?');   p.append(year)
    if month: w.append('strftime("%m",transaction_date)=?');   p.append(str(month).zfill(2))
    ws = ' AND '.join(w)
    rows = qry(f'SELECT * FROM transactions WHERE {ws} ORDER BY transaction_date DESC', p)
    ingresos = sum(r['amount'] for r in rows if r['type']=='income')
    gastos   = sum(r['amount'] for r in rows if r['type']=='expense')
    neto = ingresos - gastos
    margen = round((neto/ingresos*100),1) if ingresos>0 else 0
    # por categoria
    cats = {}
    for r in rows:
        c = r.get('category_id') or 'Sin categoría'
        cats.setdefault(c, {'income':0,'expense':0})
        cats[c][r['type'] if r['type'] in ('income','expense') else 'expense'] += r['amount']
    por_categoria = [{'categoria':k,'income':v['income'],'expense':v['expense']} for k,v in cats.items()]
    ultimos = rows[:10]
    expense_bd = {}
    income_bd = {}
    for r in rows:
        vc = r.get('vendor_client') or 'Otros'
        if r['type'] == 'expense':
            expense_bd[vc] = expense_bd.get(vc, 0) + r['amount']
        else:
            income_bd[vc] = income_bd.get(vc, 0) + r['amount']
    withdrawable = round(neto * 0.4, 2) if neto > 0 else 0
    return jsonify({
        'summary': {
            'income': ingresos, 'expenses': gastos, 'net': neto,
            'margin_pct': margen, 'withdrawable': withdrawable,
            'tx_count': len(rows)
        },
        'expense_breakdown': expense_bd,
        'income_breakdown': income_bd,
        'recent_transactions': ultimos,
        'ingresos': ingresos, 'gastos': gastos, 'neto': neto, 'margen': margen
    })


init_db()


@app.route('/api/transactions/upload', methods=['POST'])
@login_required
def upload_transaction():
    import anthropic, base64, re
    if 'file' not in request.files:
        return jsonify({'error': 'No file'}), 400
    f = request.files['file']
    tx_type = request.form.get('type', 'expense')
    empresa_id = request.form.get('empresa_id') or get_first_empresa(session['user_id'])
    data = f.read()
    b64 = base64.standard_b64encode(data).decode()
    mt = f.mimetype or 'image/jpeg'
    if mt == 'application/pdf':
        media_type = 'application/pdf'
        src = {"type": "base64", "media_type": media_type, "data": b64}
        content_block = {"type": "document", "source": src}
    else:
        media_type = mt if mt in ['image/jpeg','image/png','image/webp','image/gif'] else 'image/jpeg'
        src = {"type": "base64", "media_type": media_type, "data": b64}
        content_block = {"type": "image", "source": src}
    client = anthropic.Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY'))
    msg = client.messages.create(
        model='claude-opus-4-5',
        max_tokens=512,
        messages=[{"role":"user","content":[
            content_block,
            {"type":"text","text":"Extrae de esta factura/ticket: importe_total (número), iva (número o 0), proveedor (texto), fecha (YYYY-MM-DD), categoria (una de: alimentacion,bebidas,suministros,nominas,alquiler,marketing,otros). Responde SOLO con JSON: {importe, iva, proveedor, fecha, categoria}"}
        ]}]
    )
    raw = msg.content[0].text.strip()
    raw = re.sub(r'^```json|^```|```$', '', raw, flags=re.MULTILINE).strip()
    import json as _json
    parsed = _json.loads(raw)
    uid = session['user_id']
    exe(
        "INSERT INTO transactions (empresa_id,type,amount,tax_amount,description,vendor_client,transaction_date,payment_method,source) VALUES (?,?,?,?,?,?,?,?,?)",
        [empresa_id, tx_type, parsed.get('importe',0), parsed.get('iva',0),
         parsed.get('proveedor','Factura'), parsed.get('proveedor',''), parsed.get('fecha', str(datetime.date.today())), 'cash', 'ocr']
    )
    return jsonify({'ok': True, 'parsed': parsed})

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5001))
    print(f'\n  MATSEBIAN ERP v2.0\n  http://localhost:{port}\n  Login: daniel@matsebian.com / admin123\n')
    app.run(host='0.0.0.0', port=port, debug=True)

@app.route('/api/reports/', methods=['GET'])
@login_required
def get_reports():
    eid = get_first_empresa(session['user_id'])
    if not eid: return jsonify([])
    reports = qry('SELECT * FROM reports WHERE empresa_id=? ORDER BY created_at DESC', [eid])
    return jsonify(reports)

@app.route('/api/reports/generate/<int:year>/<int:month>', methods=['POST'])
@login_required
def generate_report(year, month):
    import anthropic
    eid = get_first_empresa(session['user_id'])
    if not eid: return jsonify({'error': 'No empresa'}), 400
    rows = qry('SELECT * FROM transactions WHERE empresa_id=? AND strftime("%Y",transaction_date)=? AND strftime("%m",transaction_date)=?',
               [eid, str(year), str(month).zfill(2)])
    ingresos = sum(r['amount'] for r in rows if r['type']=='income')
    gastos   = sum(r['amount'] for r in rows if r['type']=='expense')
    neto = ingresos - gastos
    resumen_tx = '\n'.join([f"- {r['transaction_date']} {r['type']} €{r['amount']} {r['description']}" for r in rows[:30]])
    client = anthropic.Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY'))
    msg = client.messages.create(
        model='claude-opus-4-5', max_tokens=600,
        messages=[{"role":"user","content":f"Eres asesor financiero de un restaurante/franquicia español. Datos del mes {month}/{year}:\nIngresos: €{ingresos}\nGastos: €{gastos}\nNeto: €{neto}\nMovimientos:\n{resumen_tx}\n\nEscribe un reporte mensual breve (3-4 párrafos) con: resumen de resultados, análisis, y recomendación de acción concreta. En español, directo y profesional."}]
    )
    texto = msg.content[0].text
    exe('INSERT INTO reports (empresa_id, year, month, content, created_at) VALUES (?,?,?,?,datetime("now"))',
        [eid, year, month, texto])
    return jsonify({'ok': True, 'content': texto})
